from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, request
import json
import datetime
import time

Lab2Mazlov   = 'Lab2-Mazlov'
xls_filename = 'data.xlsx'
buffer       = []
max_buffer_size = 5

def SetWorkSheet(wb): #wb, only_my_sheet=True):
    def DelOtherSheets(wb):
        sheets=wb.sheetnames
        if len(sheets)>1:
            print('   > deleting other sheets ...', end='')
        # удаляем ненужные листы, оставляем только свой
        for i in sheets:
            if i==Lab2Mazlov: continue
            wb.remove(wb[i])
        print('ok', end='')

    ListHeader = [
    ['№',       {'size':16, 'color':'f2e205', 'bold':True}, {'fill_type':'solid', 'start_color':'383636'}],
    ['User ID', {'size':16, 'color':'f2e205', 'bold':True}, {'fill_type':'solid', 'start_color':'383636'}],
    ['Datetime',{'size':16, 'color':'f2e205', 'bold':True}, {'fill_type':'solid', 'start_color':'383636'}],
    ['Item',    {'size':16, 'color':'f2e205', 'bold':True}, {'fill_type':'solid', 'start_color':'383636'}],
    ['Price',   {'size':16, 'color':'f2e205', 'bold':True}, {'fill_type':'solid', 'start_color':'383636'}]
    ]

    ws = wb.active
    sheets=wb.sheetnames
    if Lab2Mazlov not in sheets:
        print('   > adding new sheet', Lab2Mazlov, '... ok.')
        ws2 = wb.create_sheet(Lab2Mazlov, 0)
        DelOtherSheets(wb)
        for col in range(len(ListHeader)):
            d = ws2.cell(1, col+1, ListHeader[col][0])
            d.font = Font(**ListHeader[col][1])                                 #name='Arial'
            d.fill = PatternFill(**ListHeader[col][2])                          #ws2['C:D'].width = 90
            d.alignment = Alignment(horizontal='center', vertical='center' )    #wrap_text=True
    else:
        print('   > sheet', Lab2Mazlov, 'found. ok.')
        DelOtherSheets(wb)
    return wb.active  # wb.copy_worksheet(ws) ws = wb.worksheets[-1]  ws.title = city

def CheckFileReadyToWork(filename):
    print('> Checking file ... ', filename)
    try:
        wb = load_workbook(filename)    # -> FileNotFoundError
        tmp=SetWorkSheet(wb)
        wb.save(filename)               # -> PermissionError
        wb.close()

    except PermissionError as e:        # ошибка, если файл в этот момент открыт в Excel, а мы туда пишем save
        print('   > file is busy at the moment. Close it. failed.' )
        return False

    except FileNotFoundError as e:      # ошибка, если файл ещё не существует, а мы его открываем
        print('   > file not found.\n   > creating new.')
        wb = Workbook()
        tmp=SetWorkSheet(wb)
        wb.save(filename)
        print('   > file', filename, 'created. ok.')
        wb.close()
        return True

    except Exception as e:              # какая-то общая. Общий случай
        print('   > General error:', e)
        return False
    else:
        return True

def WriteToFile(filename): #xls_filename
    if not CheckFileReadyToWork(filename): return False
    wb = load_workbook(filename)
    ws = wb.active
    line_cnt = ws.max_row
    for one_line in buffer:
        ws.append([line_cnt]+one_line)
        line_cnt+=1
    wb.save("data.xlsx")
    wb.close()
    return True

app = Flask(__name__)
@app.route('/', methods=['POST'])
def index():
    global max_buffer_size, Lab2Mazlov, xls_filename, buffer

    in_data = request.json

    for line in in_data['check']:
        buffer.append( [in_data['user_id'], str(datetime.datetime.now()), line['item'], line['price'] ])
        if len(buffer) >= max_buffer_size:
            print('\n> Buffer is full. Writing to file', xls_filename, '...')

            if WriteToFile(xls_filename):
                buffer=[]
                print('\n   > Buffer has been writen. ok.')
            else:
                print('\n   > Error writing file', xls_filename)
                print('      > Buffer length +5.')
                max_buffer_size+=len(in_data['check'])

    return "Ok"

if __name__ == "__main__":
    app.run()


'''
test= { "user_id": "9359F683B13A18A1",
        "check":[ {"item": "сандальки",  "price": 250},
                  {"item": "носки",      "price": 100},
                  {"item": "носки2",     "price": 200},
                  {"item": "носки3",     "price": 300},
                  {"item": "носки4",     "price": 400},
                  {"item": "носки5",     "price": 500},
                  {"item": "носки6",     "price": 600}
                  ]   }
print('max_buffer_size=',max_buffer_size)
index(test)
print('max_buffer_size=',max_buffer_size)
print(buffer)
'''
